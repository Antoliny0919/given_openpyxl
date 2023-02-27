#Python stdlib imports
import os
import warnings
from zipfile import ZipFile, ZIP_DEFLATED, BadZipfile
from sys import exc_info

from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.reader.excel import _load_workbook, repair_central_directory
from openpyxl.compat import unicode, file
from openpyxl import Workbook

from convert.xls_convert import convert_xlsx


KEEP_VBA = os.environ.get("OPENPYXL_KEEP_VBA", "False") == "True"
CENTRAL_DIRECTORY_SIGNATURE = b'\x50\x4b\x05\x06'

# Add .xls, .xlsb extension
SUPPORTED_FORMATS = ('.xlsx', '.xlsm', '.xltx', '.xltm', '.xls', 'xlsb')

# .xls, .xlsb extenstion regenerate to xlsx extension use xlrd library
REGENERATION_FORMATS = ('.xls', 'xlsb')


def load_workbook(filename, read_only=False, use_iterators=False, keep_vba=KEEP_VBA, guess_types=False, data_only=False):
    """Open the given filename and return the workbook
    :param filename: the path to open or a file-like object
    :type filename: string or a file-like object open in binary mode c.f., :class:`zipfile.ZipFile`
    :param read_only: optimised for reading, content cannot be edited
    :type read_only: bool
    :param use_iterators: use lazy load for cells
    :type use_iterators: bool
    :param keep_vba: preseve vba content (this does NOT mean you can use it)
    :type keep_vba: bool
    :param guess_types: guess cell content type and do not read it from the file
    :type guess_types: bool
    :param data_only: controls whether cells with formulae have either the formula (default) or the value stored the last time Excel read the sheet
    :type data_only: bool
    :rtype: :class:`openpyxl.workbook.Workbook`
    .. note::
        When using lazy load, all worksheets will be :class:`openpyxl.worksheet.iter_worksheet.IterableWorksheet`
        and the returned workbook will be read-only.
    """
    read_only = read_only or use_iterators

    is_file_like = hasattr(filename, 'read')

    if not is_file_like and os.path.isfile(filename):
        file_format = os.path.splitext(filename)[-1]
        if file_format not in SUPPORTED_FORMATS:
            # if extends --> (.xlsb or xls) use win32 module convert xlsx file
            # use win32 module so only run window environment
            if file_format == '.xlsb' or  file_format == '.xls':
                msg = ('openpyxl does not support binary format .xlsb and .xls'
                    'so use win32com module convert extends xlsb and xls to xlsx'
                    'reloading please new xlsx files(convert file)')
                convert_xlsx(filename, file_format)
                
            else:
                msg = ('openpyxl does not support %s file format, '
                    'please check you can open '
                    'it with Excel first. '
                    'Supported formats are: %s') % (file_format,
                                                    ','.join(SUPPORTED_FORMATS))
                raise InvalidFileException(msg)


    if is_file_like:
        # fileobject must have been opened with 'rb' flag
        # it is required by zipfile
        if getattr(filename, 'encoding', None) is not None:
            raise IOError("File-object must be opened in binary mode")

    try:
        archive = ZipFile(filename, 'r', ZIP_DEFLATED)
    except BadZipfile:
        f = repair_central_directory(filename, is_file_like)
        archive = ZipFile(f, 'r', ZIP_DEFLATED)
        
    wb = Workbook(guess_types=guess_types, data_only=data_only, read_only=read_only)

    if read_only and guess_types:
        warnings.warn('Data types are not guessed when using iterator reader')

    try:
        _load_workbook(wb, archive, filename, read_only, keep_vba)
    except KeyError:
        e = exc_info()[1]
        raise InvalidFileException(unicode(e))

    archive.close()
    return wb